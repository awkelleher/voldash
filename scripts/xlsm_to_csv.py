"""
Extract vol/skew data from hertzsoy.XX.VOLS.xlsm workbook to flat CSV.

Reads directly from the Excel workbook (no CSV export needed) and outputs
the same format as master_vol_skew.csv for pasting into the Streamlit dashboard.

Copies to clipboard automatically for quick paste-in.

Usage:
    python scripts/xlsm_to_csv.py                                    # uses data/hertzsoy.22.VOLS.xlsm
    python scripts/xlsm_to_csv.py data/hertzsoy.22.VOLS.xlsm         # specify file
    python scripts/xlsm_to_csv.py data/hertzsoy.22.VOLS.xlsm -o out.csv  # save to file
"""

import sys
import os
import argparse
from datetime import datetime, date

import openpyxl
import pandas as pd

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ---------------------------------------------------------------------------
# Workbook layout constants
# ---------------------------------------------------------------------------

# VOLLINKS sheet: SOY, MEAL, OIL, CORN (rows are 1-indexed openpyxl)
# Each block: header row, then up to 12 data rows, then padding
VOLLINKS_BLOCKS = {
    'SOY':  {'start': 2,  'end': 13, 'code': 'S'},
    'MEAL': {'start': 27, 'end': 38, 'code': 'SM'},
    'OIL':  {'start': 52, 'end': 63, 'code': 'BO'},
    'CORN': {'start': 77, 'end': 88, 'code': 'C'},
}

# VOLLINKS column indices (1-indexed)
VL_COL = {
    'option_code': 3,   # C
    'dirty_vol':   4,   # D
    'skew_m1_5':  10,   # J  (DifferentialD)
    'skew_m0_5':  11,   # K  (DifferentialE)
    'skew_p0_5':  12,   # L  (DifferentialF)
    'skew_p1_5':  13,   # M  (DifferentialG)
    'skew_p3':    14,   # N  (DifferentialH)
    'expiry':     19,   # S  (OptionExpiration)
    'dte':        20,   # T  (DTE)
}

# Individual Parameters sheets for WHEAT and KW (same column layout)
PARAMS_BLOCKS = {
    'WHEAT': {'sheet': 'WHEAT Parameters', 'start': 2, 'end': 13, 'code': 'W'},
    'KW':    {'sheet': 'KW Parameters',    'start': 2, 'end': 13, 'code': 'KW'},
}

# Parameters sheet column indices (1-indexed) - same structure as VOLLINKS
PARAMS_COL = {
    'option_code': 4,   # D
    'dirty_vol':   5,   # E
    'skew_m1_5':  11,   # K  (DifferentialD)
    'skew_m0_5':  12,   # L  (DifferentialE)
    'skew_p0_5':  13,   # M  (DifferentialF)
    'skew_p1_5':  14,   # N  (DifferentialG)
    'skew_p3':    15,   # O  (DifferentialH)
    'expiry':     20,   # T  (OptionExpiration)
}

# Monthly&BEs sheet: forward vol diagonal positions
# Each commodity section is 13 rows (1 header + 12 data)
# The diagonal: row offset i -> column B + (i-1)
MONTHLY_BES_BLOCKS = {
    'SOY':  {'data_start': 2,  'max_rows': 12},  # rows 2-13
    'MEAL': {'data_start': 15, 'max_rows': 12},  # rows 15-26
    'OIL':  {'data_start': 28, 'max_rows': 12},  # rows 28-39
    'CORN': {'data_start': 41, 'max_rows': 12},  # rows 41-52
}

# Ladder sheet: forward vols for WHEAT and KW
# "Variance Day Vol" row within each block
LADDER_BLOCKS = {
    'WHEAT': {'contract_row': 50, 'fwd_vol_row': 52, 'start_col': 2, 'max_cols': 12},
    'KW':    {'contract_row': 62, 'fwd_vol_row': 64, 'start_col': 2, 'max_cols': 12},
}

EXCEL_EPOCH = datetime(1899, 12, 30)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_float(val):
    """Convert cell value to float, returning None for errors/blanks."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (val != val):  # NaN check
            return None
        return float(val)
    s = str(val).strip()
    if not s or s.startswith('#') or s.lower() in ('n/a', 'na', '#n/a', '#div/0!'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_expiry(val):
    """Convert cell value (datetime or Excel serial) to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    # Try Excel serial number
    try:
        serial = int(float(val))
        if 30000 < serial < 60000:
            from datetime import timedelta
            return (EXCEL_EPOCH + timedelta(days=serial)).strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        pass
    return None


def is_active_row(dirty_vol, expiry_str, dte):
    """Check if a row has valid, active data."""
    if dirty_vol is None or dirty_vol == 0:
        return False
    if expiry_str is None:
        return False
    # Skip rows with negative DTE (expired contracts)
    if dte is not None and dte < 0:
        return False
    # Skip stale dates (before 2025)
    try:
        year = int(expiry_str[:4])
        if year < 2025:
            return False
    except (ValueError, IndexError):
        return False
    return True


# ---------------------------------------------------------------------------
# Extraction functions
# ---------------------------------------------------------------------------
def extract_vollinks(ws, today_str):
    """Extract SOY, MEAL, OIL, CORN from VOLLINKS sheet."""
    records = []

    for commodity, block in VOLLINKS_BLOCKS.items():
        for row_idx in range(block['start'], block['end'] + 1):
            dirty_vol = parse_float(ws.cell(row=row_idx, column=VL_COL['dirty_vol']).value)
            expiry = parse_expiry(ws.cell(row=row_idx, column=VL_COL['expiry']).value)
            dte = parse_float(ws.cell(row=row_idx, column=VL_COL['dte']).value)

            if not is_active_row(dirty_vol, expiry, dte):
                continue

            records.append({
                'date': today_str,
                'commodity': commodity,
                'expiry': expiry,
                'dirty_vol': round(dirty_vol, 4),
                'fwd_vol': None,  # filled in later from Monthly&BEs
                'skew_m1.5': parse_float(ws.cell(row=row_idx, column=VL_COL['skew_m1_5']).value),
                'skew_m0.5': parse_float(ws.cell(row=row_idx, column=VL_COL['skew_m0_5']).value),
                'skew_p0.5': parse_float(ws.cell(row=row_idx, column=VL_COL['skew_p0_5']).value),
                'skew_p1.5': parse_float(ws.cell(row=row_idx, column=VL_COL['skew_p1_5']).value),
                'skew_p3.0': parse_float(ws.cell(row=row_idx, column=VL_COL['skew_p3']).value),
                'trading_dte': dte,
                '_position': row_idx - block['start'],  # for matching fwd_vol
            })

    return records


def extract_params_sheet(wb, commodity, block, today_str):
    """Extract WHEAT or KW from their individual Parameters sheet."""
    try:
        ws = wb[block['sheet']]
    except KeyError:
        print(f"  Warning: sheet '{block['sheet']}' not found", file=sys.stderr)
        return []

    records = []
    for row_idx in range(block['start'], block['end'] + 1):
        dirty_vol = parse_float(ws.cell(row=row_idx, column=PARAMS_COL['dirty_vol']).value)
        expiry = parse_expiry(ws.cell(row=row_idx, column=PARAMS_COL['expiry']).value)

        # No DTE column in individual Parameters sheets - compute from expiry
        dte = None
        if expiry:
            try:
                exp_date = datetime.strptime(expiry, '%Y-%m-%d').date()
                dte = (exp_date - date.today()).days
            except ValueError:
                pass

        if not is_active_row(dirty_vol, expiry, dte):
            continue

        records.append({
            'date': today_str,
            'commodity': commodity,
            'expiry': expiry,
            'dirty_vol': round(dirty_vol, 4),
            'fwd_vol': None,  # filled in later from Ladder
            'skew_m1.5': parse_float(ws.cell(row=row_idx, column=PARAMS_COL['skew_m1_5']).value),
            'skew_m0.5': parse_float(ws.cell(row=row_idx, column=PARAMS_COL['skew_m0_5']).value),
            'skew_p0.5': parse_float(ws.cell(row=row_idx, column=PARAMS_COL['skew_p0_5']).value),
            'skew_p1.5': parse_float(ws.cell(row=row_idx, column=PARAMS_COL['skew_p1_5']).value),
            'skew_p3.0': parse_float(ws.cell(row=row_idx, column=PARAMS_COL['skew_p3']).value),
            'trading_dte': dte,
            '_position': row_idx - block['start'],  # for matching fwd_vol
        })

    return records


def fill_fwd_vol_from_monthly_bes(ws, records):
    """Fill fwd_vol for SOY, MEAL, OIL, CORN from Monthly&BEs diagonal."""
    for commodity, block in MONTHLY_BES_BLOCKS.items():
        # Build list of fwd_vols from the diagonal
        fwd_vols = []
        for i in range(block['max_rows']):
            row_idx = block['data_start'] + i
            col_idx = 2 + i  # column B=2, C=3, D=4, ... (diagonal)
            val = parse_float(ws.cell(row=row_idx, column=col_idx).value)
            fwd_vols.append(val)

        # Match by position within the commodity block
        comm_records = [r for r in records if r['commodity'] == commodity]
        for rec in comm_records:
            pos = rec['_position']
            if pos < len(fwd_vols) and fwd_vols[pos] is not None:
                rec['fwd_vol'] = round(fwd_vols[pos], 4)


def fill_fwd_vol_from_ladder(ws, records):
    """Fill fwd_vol for WHEAT and KW from Ladder sheet."""
    for commodity, block in LADDER_BLOCKS.items():
        # Read fwd_vols from the "Variance Day Vol" row
        fwd_vols = []
        for col_idx in range(block['start_col'], block['start_col'] + block['max_cols']):
            val = parse_float(ws.cell(row=block['fwd_vol_row'], column=col_idx).value)
            fwd_vols.append(val)

        # Match by position
        comm_records = [r for r in records if r['commodity'] == commodity]
        for rec in comm_records:
            pos = rec['_position']
            if pos < len(fwd_vols) and fwd_vols[pos] is not None:
                rec['fwd_vol'] = round(fwd_vols[pos], 4)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def extract_from_xlsm(filepath):
    """Extract all vol/skew data from the .xlsm workbook."""
    print(f"Opening {os.path.basename(filepath)} ...", file=sys.stderr)
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    today_str = date.today().strftime('%Y-%m-%d')

    # 1. Extract SOY, MEAL, OIL, CORN from VOLLINKS
    vollinks_ws = wb['VOLLINKS']
    records = extract_vollinks(vollinks_ws, today_str)
    print(f"  VOLLINKS: {len(records)} rows (SOY/MEAL/OIL/CORN)", file=sys.stderr)

    # 2. Extract WHEAT, KW from individual Parameters sheets
    for commodity, block in PARAMS_BLOCKS.items():
        params_records = extract_params_sheet(wb, commodity, block, today_str)
        records.extend(params_records)
        print(f"  {block['sheet']}: {len(params_records)} rows", file=sys.stderr)

    # 3. Fill forward vols from Monthly&BEs (SOY/MEAL/OIL/CORN)
    monthly_ws = wb['Monthly&BEs']
    fill_fwd_vol_from_monthly_bes(monthly_ws, records)

    # 4. Fill forward vols from Ladder (WHEAT/KW)
    ladder_ws = wb['Ladder']
    fill_fwd_vol_from_ladder(ladder_ws, records)

    # Clean up temp field
    for rec in records:
        rec.pop('_position', None)

    wb.close()
    return records


def main():
    parser = argparse.ArgumentParser(description="Extract vol/skew from .xlsm to flat CSV")
    parser.add_argument("input", nargs="?", default=None,
                        help="Input .xlsm file (default: data/hertzsoy.22.VOLS.xlsm)")
    parser.add_argument("-o", "--output", default=None,
                        help="Output file (default: print to stdout + clipboard)")
    args = parser.parse_args()

    # Resolve input path
    if args.input is None:
        input_file = os.path.join(PROJECT_DIR, 'data', 'hertzsoy.22.VOLS.xlsm')
    elif os.path.isabs(args.input):
        input_file = args.input
    else:
        input_file = os.path.join(PROJECT_DIR, args.input)

    if not os.path.exists(input_file):
        print(f"File not found: {input_file}", file=sys.stderr)
        sys.exit(1)

    # Extract
    records = extract_from_xlsm(input_file)
    if not records:
        print("No records extracted.", file=sys.stderr)
        sys.exit(1)

    df = pd.DataFrame(records)

    # Deduplicate (some sheets have duplicate rows, e.g. KW)
    # Keep the first occurrence (which has fwd_vol if available)
    df = df.sort_values('fwd_vol', ascending=False, na_position='last')
    df = df.drop_duplicates(subset=['date', 'commodity', 'expiry'], keep='first')

    # Order columns to match master_vol_skew.csv
    col_order = ['date', 'commodity', 'expiry', 'dirty_vol', 'fwd_vol',
                 'skew_m1.5', 'skew_m0.5', 'skew_p0.5', 'skew_p1.5', 'skew_p3.0',
                 'trading_dte']
    df = df[[c for c in col_order if c in df.columns]]
    df = df.sort_values(['commodity', 'expiry'], ascending=[True, True])

    # Output
    csv_text = df.to_csv(index=False)

    if args.output:
        df.to_csv(args.output, index=False)
        print(f"Saved {len(df)} rows to {args.output}", file=sys.stderr)
    else:
        print(csv_text, end='')

    # Copy to clipboard (Windows)
    try:
        import subprocess
        proc = subprocess.Popen(['clip'], stdin=subprocess.PIPE)
        proc.communicate(csv_text.encode('utf-8'))
        print(f"\n({len(df)} rows copied to clipboard)", file=sys.stderr)
    except Exception:
        pass

    # Summary
    print(f"\nExtracted {len(df)} rows for {df['commodity'].nunique()} commodities", file=sys.stderr)
    for comm in ['SOY', 'MEAL', 'OIL', 'CORN', 'WHEAT', 'KW']:
        comm_df = df[df['commodity'] == comm]
        if not comm_df.empty:
            fwd_count = comm_df['fwd_vol'].notna().sum()
            print(f"  {comm}: {len(comm_df)} contracts, {fwd_count} with fwd_vol", file=sys.stderr)


if __name__ == '__main__':
    main()
