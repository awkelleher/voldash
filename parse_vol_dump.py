"""
Vol/Skew Data Parser
====================
Parses the wide-format vol dump (CSV or XLSM) into a flat master CSV.

Usage:
  python parse_vol_dump.py                          # uses defaults (eod_vol_snap.csv -> master_vol_skew.csv)
  python parse_vol_dump.py -i other_dump.csv        # custom input
  python parse_vol_dump.py -i dump1.csv dump2.csv   # multiple inputs
  python parse_vol_dump.py -o my_master.csv         # custom output
"""

import argparse
import csv
import logging
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import pandas as pd

# --- Config ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT = os.path.join(SCRIPT_DIR, 'eod_vol_snap.csv')
DEFAULT_OUTPUT = os.path.join(SCRIPT_DIR, 'master_vol_skew.csv')
LOG_FILE = os.path.join(SCRIPT_DIR, 'vol_parser.log')

EXCEL_EPOCH = datetime(1899, 12, 30)

COMMODITY_BLOCKS = {
    'SOY':   {'row_num': 1,   'expiry': 2,   'dirty_vol': 4,   'skew_m1_5': 11,  'skew_m0_5': 12,  'skew_p0_5': 13,  'skew_p1_5': 14,  'skew_p3': 15,  'trading_dte': 16},
    'MEAL':  {'row_num': 19,  'expiry': 20,  'dirty_vol': 22,  'skew_m1_5': 29,  'skew_m0_5': 30,  'skew_p0_5': 31,  'skew_p1_5': 32,  'skew_p3': 33,  'trading_dte': 34},
    'OIL':   {'row_num': 37,  'expiry': 38,  'dirty_vol': 40,  'skew_m1_5': 47,  'skew_m0_5': 48,  'skew_p0_5': 49,  'skew_p1_5': 50,  'skew_p3': 51,  'trading_dte': 52},
    'CORN':  {'row_num': 55,  'expiry': 56,  'dirty_vol': 58,  'skew_m1_5': 65,  'skew_m0_5': 66,  'skew_p0_5': 67,  'skew_p1_5': 68,  'skew_p3': 69,  'trading_dte': 70},
    'WHEAT': {'row_num': 83,  'expiry': 84,  'dirty_vol': 86,  'skew_m1_5': 93,  'skew_m0_5': 94,  'skew_p0_5': 95,  'skew_p1_5': 96,  'skew_p3': 97,  'trading_dte': 98},
    'KW':    {'row_num': 101, 'expiry': 102, 'dirty_vol': 104, 'skew_m1_5': 111, 'skew_m0_5': 112, 'skew_p0_5': 113, 'skew_p1_5': 114, 'skew_p3': 115, 'trading_dte': 116},
}

MAX_EXPIRIES_PER_BLOCK = 12


def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s  %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        handlers=[
            logging.FileHandler(LOG_FILE),
            logging.StreamHandler(sys.stdout),
        ]
    )


def serial_to_date(serial):
    try:
        s = int(float(serial))
        if 30000 < s < 60000:
            return (EXCEL_EPOCH + timedelta(days=s)).strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        pass
    return None


def parse_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (val != val):
            return None
        return float(val)
    val = str(val).strip()
    if not val or val.startswith('#') or val.lower() in ('n/a', 'na'):
        return None
    try:
        return float(val)
    except ValueError:
        return None


def parse_csv_file(filepath):
    records = []
    current_date = None

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 20:
                continue

            col0 = row[0].strip() if row[0] else ''
            col1 = row[1].strip() if len(row) > 1 else ''

            if col0 and col1 == 'SOY':
                current_date = serial_to_date(col0)
                continue

            if current_date is None:
                continue

            row_num = parse_float(col1)
            if row_num is None or row_num < 1 or row_num > 15:
                continue

            for commodity, cols in COMMODITY_BLOCKS.items():
                try:
                    expiry_raw = row[cols['expiry']] if cols['expiry'] < len(row) else ''
                    dirty_vol = parse_float(row[cols['dirty_vol']] if cols['dirty_vol'] < len(row) else '')

                    expiry = serial_to_date(expiry_raw)
                    if expiry is None or dirty_vol is None or dirty_vol == 0:
                        continue

                    records.append({
                        'date': current_date,
                        'commodity': commodity,
                        'expiry': expiry,
                        'dirty_vol': round(dirty_vol, 4),
                        'skew_m1.5': parse_float(row[cols['skew_m1_5']] if cols['skew_m1_5'] < len(row) else ''),
                        'skew_m0.5': parse_float(row[cols['skew_m0_5']] if cols['skew_m0_5'] < len(row) else ''),
                        'skew_p0.5': parse_float(row[cols['skew_p0_5']] if cols['skew_p0_5'] < len(row) else ''),
                        'skew_p1.5': parse_float(row[cols['skew_p1_5']] if cols['skew_p1_5'] < len(row) else ''),
                        'skew_p3.0': parse_float(row[cols['skew_p3']] if cols['skew_p3'] < len(row) else ''),
                        'trading_dte': parse_float(row[cols['trading_dte']] if cols['trading_dte'] < len(row) else ''),
                    })
                except (IndexError, KeyError):
                    continue

    return records


def parse_xlsm_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value is not None and isinstance(cell.value, datetime):
            header_rows.append((cell.row, cell.value))

    records = []
    for header_row, trade_date in header_rows:
        trade_date_str = trade_date.strftime('%Y-%m-%d')
        for data_row in range(header_row + 1, header_row + 1 + MAX_EXPIRIES_PER_BLOCK):
            if data_row > ws.max_row:
                break
            for commodity, cols in COMMODITY_BLOCKS.items():
                row_num_val = ws.cell(row=data_row, column=cols['row_num'] + 1).value
                if not isinstance(row_num_val, (int, float)) or row_num_val < 1:
                    continue

                expiry_val = ws.cell(row=data_row, column=cols['expiry'] + 1).value
                dirty_vol = parse_float(ws.cell(row=data_row, column=cols['dirty_vol'] + 1).value)

                if not isinstance(expiry_val, datetime):
                    continue
                if dirty_vol is None or dirty_vol == 0:
                    continue

                def cell_val(col_idx):
                    return parse_float(ws.cell(row=data_row, column=col_idx + 1).value)

                records.append({
                    'date': trade_date_str,
                    'commodity': commodity,
                    'expiry': expiry_val.strftime('%Y-%m-%d'),
                    'dirty_vol': round(dirty_vol, 4),
                    'skew_m1.5': cell_val(cols['skew_m1_5']),
                    'skew_m0.5': cell_val(cols['skew_m0_5']),
                    'skew_p0.5': cell_val(cols['skew_p0_5']),
                    'skew_p1.5': cell_val(cols['skew_p1_5']),
                    'skew_p3.0': cell_val(cols['skew_p3']),
                    'trading_dte': cell_val(cols['trading_dte']),
                })

    wb.close()
    return records


def parse_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        return parse_xlsm_file(filepath)
    else:
        return parse_csv_file(filepath)


def main():
    setup_logging()
    log = logging.getLogger()

    parser = argparse.ArgumentParser(description='Parse vol/skew dump to flat CSV')
    parser.add_argument('-i', '--input', nargs='+', default=[DEFAULT_INPUT], help=f'Input file(s) (default: {DEFAULT_INPUT})')
    parser.add_argument('-o', '--output', default=DEFAULT_OUTPUT, help=f'Output CSV (default: {DEFAULT_OUTPUT})')
    args = parser.parse_args()

    log.info('=' * 60)
    log.info('Vol parser started')

    # Validate inputs exist
    for filepath in args.input:
        if not os.path.exists(filepath):
            log.error(f'Input file not found: {filepath}')
            sys.exit(1)

    # Parse all input files
    all_records = []
    for filepath in args.input:
        log.info(f'Parsing {os.path.basename(filepath)}...')
        records = parse_file(filepath)
        log.info(f'  {len(records)} records extracted')
        all_records.extend(records)

    new_df = pd.DataFrame(all_records)
    if new_df.empty:
        log.warning('No valid data found in input. Exiting.')
        sys.exit(0)

    new_dates = sorted(new_df['date'].unique())
    log.info(f'New data: {len(new_df)} records, {len(new_dates)} dates ({new_dates[0]} to {new_dates[-1]})')

    # Merge with existing master
    if os.path.exists(args.output):
        existing_df = pd.read_csv(args.output)
        before_count = len(existing_df)
        log.info(f'Existing master: {before_count} records ({existing_df["date"].min()} to {existing_df["date"].max()})')

        combined = pd.concat([existing_df, new_df], ignore_index=True)
        combined = combined.drop_duplicates(subset=['date', 'commodity', 'expiry'], keep='last')
        added = len(combined) - before_count
        log.info(f'Merged: {added} new records added, {len(new_df) - added} duplicates skipped')
    else:
        combined = new_df.drop_duplicates(subset=['date', 'commodity', 'expiry'], keep='last')
        log.info(f'Created new master with {len(combined)} records')

    combined = combined.sort_values(['commodity', 'date', 'expiry']).reset_index(drop=True)
    combined.to_csv(args.output, index=False)
    log.info(f'Saved {len(combined)} total records to {os.path.basename(args.output)}')
    log.info(f'Date range: {combined["date"].min()} to {combined["date"].max()}')
    log.info('Done.')


if __name__ == '__main__':
    main()
